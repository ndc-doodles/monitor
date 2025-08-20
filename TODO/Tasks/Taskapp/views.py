from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.hashers import make_password, check_password
from .models import Register, Task
from .forms import TaskForm
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib import messages
from datetime import date, datetime, timedelta
from . import *
from django.urls import reverse

from datetime import date, datetime
import os
import io
from django.http import HttpResponse
import pandas as pd
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_protect

from .models import *
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import Register, DailyPlan

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import *  # adjust as needed
from io import BytesIO
from openpyxl import Workbook
from django.contrib.auth.hashers import check_password, make_password

# def index(request):
#     username = request.session.get('username')
#     user_id = request.session.get('user_id')

#     if not username or not user_id:
#         return redirect('login')

#     user = Register.objects.get(id=user_id)
#     tasks = Task.objects.filter(user=user)

#     if request.method == 'POST':
#         form = TaskForm(request.POST)
#         if form.is_valid():
#             task = form.save(commit=False)
#             task.user = user
#             task.save()
#             return redirect('index')
#     else:
#         form = TaskForm()

#     context = {
#         'username': username,
#         'tasks': tasks,
#         'form': form,
#     }
#     return render(request, 'index.html', context)

def AdminLogins(request):
    
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm = request.POST.get('password2')

        if password != confirm:
            return render(request, 'register.html', {'error': 'Passwords do not match'})

        if AdminLogin.objects.filter(username=username).exists():
            return render(request, 'admin_register.html', {'error': 'Username already exists'})

        hashed_password = make_password(password)
        users = AdminLogin.objects.create(username=username, password=hashed_password)
        request.session['user_id'] = users.id
        request.session['username'] = users.username
        return redirect('admin_index')

    return render(request, 'admin_register.html')

# from django.conf import settings



# def AdminLogins(request):
#     print("DB in view:", settings.DATABASES['default']['NAME'])  # Debug
#     emp = Register.objects.all()
#     emp_count = emp.count()
#     print("EMP COUNT:", emp_count)

#     if request.method == 'POST':
#         username = request.POST.get('username')
#         password = request.POST.get('password')
#         confirm = request.POST.get('password2')

#         if password != confirm:
#             return render(request, 'register.html', {'error': 'Passwords do not match'})

#         if AdminLogin.objects.filter(username=username).exists():
#             return render(request, 'admin_register.html', {'error': 'Username already exists'})

#         hashed_password = make_password(password)
#         users = AdminLogin.objects.create(username=username, password=hashed_password)
#         request.session['user_id'] = users.id
#         request.session['username'] = users.username
#         return redirect('admin_index')

#     context = {
#         'emp': emp,
#         'emp_count': emp_count,
#         'db_path': settings.DATABASES['default']['NAME']  # Optional, for debugging in template
#     }
#     return render(request, 'admin_register.html', context)


def admin_login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            users = AdminLogin.objects.get(username=username)
            if check_password(password, users.password):
                request.session['user_id'] = users.id
                request.session['username'] = users.username
                return redirect('admin_index')
            else:
                return render(request, 'admin_login.html', {'error': 'Incorrect password'})
        except AdminLogin.DoesNotExist:
            return render(request, 'admin_login.html', {'error': 'User not found'})

    return render(request, 'admin_login.html')

def admin_logout_view(request):
    request.session.flush()
    return redirect('admin_login')


from django.shortcuts import redirect, get_object_or_404

def delete_employee(request, user_id):
    employee = get_object_or_404(Register, id=user_id)
    employee.delete()
    return redirect('admin_index')  # Redirect back to admin page


def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm = request.POST.get('password2')

        if password != confirm:
            return render(request, 'register.html', {'error': 'Passwords do not match'})

        if Register.objects.filter(username=username).exists():
            return render(request, 'register.html', {'error': 'Username already exists'})

        hashed_password = make_password(password)
        user = Register.objects.create(username=username, password=hashed_password)
        request.session['user_id'] = user.id
        request.session['username'] = user.username
        return redirect('index')

    return render(request, 'register.html')



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = Register.objects.get(username=username)
            if check_password(password, user.password):
                request.session['user_id'] = user.id
                request.session['username'] = user.username
                return redirect('index')
            else:
                return render(request, 'login.html', {'error': 'Incorrect password'})
        except Register.DoesNotExist:
            return render(request, 'login.html', {'error': 'User not found'})

    return render(request, 'login.html')



def logout_view(request):
    request.session.flush()
    return redirect('login')


def change_password_view(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        user_id = request.session.get('user_id')
        if not user_id:
            return redirect('login')  # Redirect if not logged in

        user = get_object_or_404(Register, id=user_id)

        # Verify old password
        if not check_password(old_password, user.password):
            return render(request, 'change_password.html', {'error': 'Old password is incorrect'})

        # Verify match
        if new_password != confirm_password:
            return render(request, 'change_password.html', {'error': 'New passwords do not match'})

        # Save new password
        user.password = make_password(new_password)
        user.save()

        return redirect('index')  # Redirect after success

    return render(request, 'change_password.html')



# def index(request):
#     user = Register.objects.get(username=request.session['username'])

#     if request.method == 'POST':
#         task_id = request.POST.get('task_id')
#         if task_id:
#             task = Task.objects.get(id=task_id)
#         else:
#             task = Task(created_by=user)
#         task.title = request.POST.get('title')
#         task.description = request.POST.get('description')
#         task.start_date = request.POST.get('start_date')
#         task.end_date = request.POST.get('end_date')
#         task.status = request.POST.get('status')
#         task.remarks = request.POST.get('remarks')
#         task.save()
#         return redirect('index')

#     # Status filter
#     status_filter = request.GET.get('status', 'all')
#     if status_filter == 'pending':
#         tasks = Task.objects.filter(created_by=user, status__iexact='Pending')
#     elif status_filter == 'inprogress':
#         tasks = Task.objects.filter(created_by=user, status__iexact='In Progress')
#     elif status_filter == 'completed':
#         tasks = Task.objects.filter(created_by=user, status__iexact='Completed')
#     else:
#         tasks = Task.objects.filter(created_by=user)

#     # Reminder logic
#     today = date.today()
#     reminder_tasks = Task.objects.filter(
#         created_by=user,
#         end_date__range=(today, today + timedelta(days=5)),
#         status__in=["Pending", "In Progress"]
#     )

#     return render(request, 'index.html', {
#         'tasks': tasks,
#         'username': user.username,
#         'reminder_tasks': reminder_tasks
#     })

def index(request):
    username = request.session.get('username')
    if not username:
        return redirect('login')  # User not logged in

    try:
        user = Register.objects.get(username=username)
    except Register.DoesNotExist:
        request.session.flush()  # Clear invalid session
        return redirect('login')

    if request.method == 'POST':
        task_id = request.POST.get('task_id')
        if task_id:
            task = Task.objects.get(id=task_id)
        else:
            task = Task(created_by=user)
        task.title = request.POST.get('title')
        task.description = request.POST.get('description')
        task.start_date = request.POST.get('start_date')
        task.end_date = request.POST.get('end_date')
        task.status = request.POST.get('status')
        task.remarks = request.POST.get('remarks')
        task.save()
        return redirect('index')

    status_filter = request.GET.get('status', 'all')
    if status_filter == 'pending':
        tasks = Task.objects.filter(created_by=user, status__iexact='Pending')
    elif status_filter == 'inprogress':
        tasks = Task.objects.filter(created_by=user, status__iexact='In Progress')
    elif status_filter == 'completed':
        tasks = Task.objects.filter(created_by=user, status__iexact='Completed')
    else:
        tasks = Task.objects.filter(created_by=user)

    today = date.today()
    reminder_tasks = Task.objects.filter(
        created_by=user,
        end_date__range=(today, today + timedelta(days=5)),
        status__in=["Pending", "In Progress"]
    )

    return render(request, 'index.html', {
        'tasks': tasks,
        'username': user.username,
        'reminder_tasks': reminder_tasks
    })



@csrf_exempt
def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    if request.method == 'POST':
        task.title = request.POST.get('title')
        task.description = request.POST.get('description')
        task.start_date = request.POST.get('start_date')
        task.end_date = request.POST.get('end_date')
        task.status = request.POST.get('status')
        task.remarks = request.POST.get('remarks')
        task.save()
        return redirect('index')

    return JsonResponse({
        'id': task.id,
        'title': task.title,
        'description': task.description,
        'start_date': str(task.start_date),
        'end_date': str(task.end_date),
        'status': task.status,
        'remarks': task.remarks,
    })

@csrf_exempt
def delete_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.delete()
    return redirect('index')







EXCEL_FILE = os.path.join(settings.BASE_DIR, 'daily_plans.xlsx')

def daily_plan_view(request):
    # ✅ Require login
    if 'user_id' not in request.session:
        return redirect('login')

    user = get_object_or_404(Register, id=request.session['user_id'])

    if request.method == "POST":
        selected_date = request.POST.get("date", date.today().strftime("%Y-%m-%d"))

        excel_rows = []

        # Add multiple new plans
        for key, value in request.POST.items():
            if key.startswith("point_text_morning_"):
                idx = key.split("_")[-1]
                status = request.POST.get(f"status_morning_{idx}", "Pending")
                if value.strip():
                    DailyPlan.objects.create(
                        user=user,
                        date=selected_date,
                        session="Morning",
                        point_text=value.strip(),
                        status=status
                    )
                    excel_rows.append({
                        "User": user.username,
                        "Date": selected_date,
                        "Session": "Morning",
                        "Point Text": value.strip(),
                        "Status": status
                    })

            elif key.startswith("point_text_afternoon_"):
                idx = key.split("_")[-1]
                status = request.POST.get(f"status_afternoon_{idx}", "Pending")
                if value.strip():
                    DailyPlan.objects.create(
                        user=user,
                        date=selected_date,
                        session="Afternoon",
                        point_text=value.strip(),
                        status=status
                    )
                    excel_rows.append({
                        "User": user.username,
                        "Date": selected_date,
                        "Session": "Afternoon",
                        "Point Text": value.strip(),
                        "Status": status
                    })

        # Update existing plan (inline form)
        if "plan_id" in request.POST:
            plan = get_object_or_404(DailyPlan, id=request.POST.get("plan_id"), user=user)
            plan.session = request.POST.get("session")
            plan.point_text = request.POST.get("point_text")
            plan.status = request.POST.get("status", "Pending")
            plan.save()

        # ✅ Save to Excel
        if excel_rows:
            if os.path.exists(EXCEL_FILE):
                df_existing = pd.read_excel(EXCEL_FILE)
                df_new = pd.DataFrame(excel_rows)
                df_final = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_final = pd.DataFrame(excel_rows)

            df_final.to_excel(EXCEL_FILE, index=False)

        return redirect(f"{request.path}?date={selected_date}")

    else:
        selected_date = request.GET.get("date", date.today().strftime("%Y-%m-%d"))
        plans = DailyPlan.objects.filter(user=user, date=selected_date).order_by("session", "id")
        return render(request, "report.html", {
            "plans": plans,
            "date": selected_date
        })


def delete_plan(request, plan_id):
    if 'user_id' not in request.session:
        return redirect('login')

    user = get_object_or_404(Register, id=request.session['user_id'])
    plan = get_object_or_404(DailyPlan, id=plan_id, user=user)
    selected_date = plan.date
    plan.delete()
    return redirect(f"/daily-plan/?date={selected_date}")


# ✅ View to download Excel
from django.http import FileResponse, Http404


def download_excel(request):
    if 'user_id' not in request.session:
        return redirect('login')

    user = get_object_or_404(Register, id=request.session['user_id'])
    selected_date = request.GET.get('date')
    if not selected_date:
        return HttpResponse("Date parameter missing", status=400)

    # ✅ Get all plans for the date
    plans = DailyPlan.objects.filter(user=user, date=selected_date).order_by('session', 'id')

    # Group tasks by session
    grouped_data = {}
    for plan in plans:
        sess = plan.session
        if sess not in grouped_data:
            grouped_data[sess] = {"points": [], "statuses": []}
        grouped_data[sess]["points"].append(plan.point_text)
        grouped_data[sess]["statuses"].append(plan.status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Plans"

    # Styles
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1, value=selected_date)
    title_cell.fill = blue_fill
    title_cell.font = white_font
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 25

    # Header row
    headers = ["date", "username", "session", "point_text", "status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # Data rows (one row per session)
    start_row = 3
    sessions_order = ["Morning", "Afternoon"]  # Keep consistent order
    for i, sess in enumerate(sessions_order, start=start_row):
        if sess in grouped_data:
            points_text = "\n".join(grouped_data[sess]["points"])
            status_text = "\n".join(grouped_data[sess]["statuses"])

            # Date (merged across both sessions if first row)
            if sess == "Morning":
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + len(sessions_order) - 1, end_column=1)
                date_cell = ws.cell(row=start_row, column=1, value=selected_date)
                date_cell.alignment = center_alignment
                date_cell.border = thin_border

                ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + len(sessions_order) - 1, end_column=2)
                user_cell = ws.cell(row=start_row, column=2, value=user.username)
                user_cell.alignment = center_alignment
                user_cell.border = thin_border

            # Session
            ws.cell(row=i, column=3, value=sess).alignment = center_alignment
            ws.cell(row=i, column=3).border = thin_border

            # Points
            ws.cell(row=i, column=4, value=points_text).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=4).border = thin_border

            # Statuses
            ws.cell(row=i, column=5, value=status_text).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=5).border = thin_border

            ws.row_dimensions[i].height = max(40, 15 * len(grouped_data[sess]["points"]))

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=daily_plan_{selected_date}.xlsx'
    wb.save(response)
    return response


def download_all_users_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    selected_date = request.GET.get("date")
    if not selected_date:
        return HttpResponse("Date parameter is required", status=400)

    users = Register.objects.filter(plans__date=selected_date).distinct()

    wb = Workbook()
    ws = wb.active

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["date", "username", "session", "point_text", "status"]

    # Top merged date header row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=selected_date).alignment = center_alignment
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill

    # Second row: column headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    current_row = 3

    # Merge date across all rows for the day
    total_rows_for_day = sum(
        DailyPlan.objects.filter(user=user, date=selected_date).values_list('session', flat=True).distinct().count()
        for user in users
    )
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + total_rows_for_day - 1, end_column=1)
    ws.cell(row=current_row, column=1, value=selected_date).alignment = center_alignment
    ws.cell(row=current_row, column=1).border = thin_border

    # Loop through each user
    for user in users:
        plans = DailyPlan.objects.filter(user=user, date=selected_date).order_by("session")

        session_data = {"Morning": {"points": [], "status": []}, "Afternoon": {"points": [], "status": []}}
        for plan in plans:
            session_data.setdefault(plan.session, {"points": [], "status": []})
            session_data[plan.session]["points"].append(plan.point_text)
            session_data[plan.session]["status"].append(plan.status)

        # Keep only sessions with data
        session_data = {k: v for k, v in session_data.items() if v["points"]}

        user_total_rows = len(session_data)

        # Merge username for the user's rows
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + user_total_rows - 1, end_column=2)
        ws.cell(row=current_row, column=2, value=user.username).alignment = center_alignment
        ws.cell(row=current_row, column=2).border = thin_border

        # Place Morning first, then Afternoon
        for session_name in ["Morning", "Afternoon"]:
            if session_name in session_data:
                data = session_data[session_name]
                ws.cell(row=current_row, column=3, value=session_name).alignment = center_alignment
                ws.cell(row=current_row, column=4, value="\n".join(data["points"])).alignment = center_alignment
                ws.cell(row=current_row, column=5, value="\n".join(data["status"])).alignment = center_alignment

                for col in range(1, 6):
                    ws.cell(row=current_row, column=col).border = thin_border

                current_row += 1

    # Adjust column widths dynamically (wider for point_text and status)
    col_widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            if cell.value:
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), len(str(cell.value)) + 2)
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = min(width, 50)  # cap at 50

    # Return file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="daily_plan_all_users.xlsx"'
    wb.save(response)
    return response



def admin(request):
    emp = Register.objects.all()
    emp_count = emp.count()  # Count employees
    return render(request, 'admin_index.html', {
        'emp': emp,
        'emp_count': emp_count
    })

